#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera NYC_2026_Planificador.xlsx — planilla editable JP/Thais."""
import os
_R = os.path.dirname(os.path.abspath(__file__))   # todo se resuelve relativo a este archivo, no a donde se corra
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule

places = json.load(open("data/places.json"))["places"]
itin = json.load(open("data/itinerary.json"))["days"]
extras = json.load(open("data/extras.json"))

F = "Arial"
NAVY = "1F3864"; SKY = "DDEBF7"; YELLOW = "FFF2CC"; GREEN = "E2EFDA"
GREY = "F2F2F2"; ORANGE = "FCE4D6"
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def title(ws, text, sub="", span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=F, bold=True, size=16, color=NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        c = ws.cell(row=2, column=1, value=sub)
        c.font = Font(name=F, size=9, italic=True, color="595959")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

wb = Workbook()

# ══════════════════════════ HOJA 1: LEEME ══════════════════════════
ws = wb.active; ws.title = "LÉEME"
title(ws, "NYC · 29 agosto – 6 septiembre 2026", "Juan Pablo & Thais — planificador editable", span=6)
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 95

rows = [
 ("", "", ""),
 ("SECCIÓN", "HOJA", "QUÉ HACER ACÁ"),
 ("1", "Lugares", "★ LA HOJA PRINCIPAL. 132 lugares con toda la info. Editá SOLO las columnas amarillas: 'JP' y 'Thais' (0/1/2) y la última, 'IMPRESCINDIBLE' (SÍ, NO, o vacía si todavía no lo decidís). La columna 'En itinerario (día)' marca en verde qué lugares están agendados y en qué día. Los totales de las otras hojas se recalculan solos."),
 ("", "Columna 'Tipo'", "Subcategoría fina: Observatorio pago, Rooftop bar, Mirador gratis, Club de jazz, Deli y pastrami, Barrio étnico, etc. 40 tipos en total. Usá el filtro del encabezado para aislar uno."),
 ("2", "Itinerario", "El plan día por día, con horarios. Editá la columna 'Estado' y 'Mis notas'. La columna 'Quién' dice si el bloque es de los dos o solo de JP."),
 ("3", "Costos", "Suma automática de lo que marcaron con interés 2. Cambien los valores de 'Lugares' y esta hoja se actualiza."),
 ("4", "Transporte", "EWR ↔ Manhattan, New Jersey ↔ NYC y el subte. Con precios, horarios y veredictos."),
 ("5", "Pases", "El análisis de si conviene NY Pass / CityPASS / Go City o tickets sueltos. Con los números."),
 ("6", "Reservas", "Checklist de qué hay que reservar y cuándo. Ordenado por urgencia. Editá la columna 'Hecho'."),
 ("7", "Decisiones", "Las cosas que todavía hay que definir, con los números para decidir."),
 ("", "", ""),
 ("CÓMO LEER", "", ""),
 ("", "Celdas AMARILLAS", "Son las que ustedes editan. El resto se calcula o es información."),
 ("", "Columna 'Interés'", "0 = no · 1 = quizás · 2 = sí. Por defecto puse a Thais más bajo en lo mainstream, asumiendo que ya lo conoce. Cambienlo libremente."),
 ("", "Columna 'Turístico'", "SÍ = está en todas las guías. NO = fuera del circuito. Sirve para filtrar lo que Thais probablemente no quiera repetir."),
 ("", "Columna 'Horario de apertura'", "Horario real de cada día de la semana, para los lugares donde llegar a la hora equivocada arruina la visita. Verificado en fuentes oficiales el 16-ago-2026."),
 ("", "Columna 'Cierra'", "Día de cierre semanal. Crítico para no ir al MET un miércoles."),
 ("", "Columna 'Gratis'", "Ventanas de entrada libre o pay-what-you-wish. Acá está la mayor parte del ahorro real del viaje."),
 ("", "Coordenadas", "Las coordenadas del mapa son APROXIMADAS (~50-100 m). La dirección de la columna 'Dirección' es la buena: la app HTML abre Google Maps buscando por nombre + dirección."),
 ("", "", ""),
 ("ADVERTENCIA", "", ""),
 ("", "Datos verificados", "Todo lo marcado con precio y horario viene de fuentes oficiales consultadas el 16 de agosto de 2026. Aun así, verifiquen precios y horarios antes de cada visita: cambian."),
 ("", "Datos NO confirmados", "Están marcados con 'VERIFICAR' en el texto. Los principales: precio 2026 del Guggenheim y del Noguchi, precios de Circle Line, tabla por tramo de Go City, y las fechas del Panorama del Caribbean Carnival."),
 ("", "Lista de Google Maps", "PENDIENTE de incorporar — no pude leerla porque la extensión de Chrome no conectó."),
]
r = 4
for a, b_, c_ in rows:
    ws.cell(row=r, column=1, value=a).font = Font(name=F, bold=True, size=10, color=NAVY)
    cb = ws.cell(row=r, column=2, value=b_); cb.font = Font(name=F, bold=True, size=10)
    cc = ws.cell(row=r, column=3, value=c_); cc.font = Font(name=F, size=10)
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("SECCIÓN",) or b_ == "" and a in ("CÓMO LEER", "ADVERTENCIA"):
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SKY)
            ws.cell(row=r, column=col).font = Font(name=F, bold=True, size=10, color=NAVY)
    if a in ("CÓMO LEER", "ADVERTENCIA"):
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SKY)
    ws.row_dimensions[r].height = 30 if len(c_) > 90 else 16
    r += 1

# ══════════════════════════ HOJA 2: LUGARES ══════════════════════════
ws = wb.create_sheet("Lugares")
title(ws, "Lugares candidatos", "Editá SOLO las columnas amarillas JP y Thais · 0 = no · 1 = quizás · 2 = sí · La última columna dice en qué DÍA del itinerario está cada lugar agendado", span=22)
H = ["Lugar", "Tipo", "Categoría", "Barrio", "Borough", "JP", "Thais", "Ambos?", "US$ p/p", "Horas",
     "Turístico", "Horario de apertura", "Cierra", "Ventana gratis / PWYW", "Reserva", "Subte", "Por qué vale la pena", "Dirección",
     "En itinerario (día)", "Google Maps", "Origen", "IMPRESCINDIBLE"]
W = [34, 22, 12, 20, 11, 6, 7, 9, 9, 7, 10, 46, 16, 34, 24, 30, 78, 34, 24, 15, 16, 17]

from urllib.parse import quote
def gmaps_url(p):
    return "https://www.google.com/maps/search/?api=1&query=" + quote(
        p["n"] + (", " + p["addr"] if p.get("addr") else "") + ", New York")

# mapa pid -> días del itinerario en que aparece
_DOW3 = {"Lunes":"lun","Martes":"mar","Miércoles":"mié","Jueves":"jue","Viernes":"vie","Sábado":"sáb","Domingo":"dom"}
DAYMAP = {}
for _d in itin:
    _seen = set()
    for _b in _d["blocks"]:
        _pid = _b.get("pid")
        if not _pid or _pid in _seen: continue
        _seen.add(_pid)
        DAYMAP.setdefault(_pid, []).append(
            f'Día {_d["n"]} · {_DOW3.get(_d["dow"], _d["dow"][:3])} {int(_d["date"][8:])}/{int(_d["date"][5:7])}')
hdr(ws, 4, H, W)
ws.freeze_panes = "A5"

DIAS = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]
def horario_semana(p):
    """Texto compacto del horario semanal, agrupando días iguales."""
    hrs = p.get("hrs")
    if not hrs: return "—"
    tramos, ini = [], 0
    for i in range(1, 7):
        if hrs[i] != hrs[i-1]:
            tramos.append((ini, i-1)); ini = i
    tramos.append((ini, 6))
    partes = []
    for a, b_ in tramos:
        w = hrs[a]
        rango = DIAS[a] if a == b_ else f"{DIAS[a]} a {DIAS[b_]}"
        partes.append(f"{rango}: {w[0]}–{w[1]}" if w else f"{rango}: cerrado")
    return " · ".join(partes)

CATN = {"mirador": "Mirador", "museo": "Museo", "barrio": "Barrio", "comida": "Comida",
        "musica": "Música", "arq": "Arquitectura", "parque": "Parque", "evento": "Evento",
        "teatro": "Teatro", "transporte": "Logística"}

places_sorted = sorted(places, key=lambda p: (list(CATN).index(p["cat"]), p["typ"], p["n"]))
r = 5
for p in places_sorted:
    ws.cell(row=r, column=1, value=p["n"]).font = Font(name=F, size=10, bold=True)
    ws.cell(row=r, column=2, value=p["typ"]).font = Font(name=F, size=10, bold=True, color="1F3864")
    ws.cell(row=r, column=3, value=CATN.get(p["cat"], p["cat"]))
    ws.cell(row=r, column=4, value=p["hood"])
    ws.cell(row=r, column=5, value=p["boro"])
    for col, val in ((6, p["jp"]), (7, p["th"])):
        c = ws.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.font = Font(name=F, size=10, bold=True, color="0000FF")
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=8, value=f'=IF(AND(F{r}=2,G{r}=2),"AMBOS",IF(AND(F{r}=2,G{r}<2),"solo JP",IF(AND(F{r}<2,G{r}=2),"solo Thais","-")))')
    c = ws.cell(row=r, column=9, value=p["cost"]); c.number_format = '$#,##0;($#,##0);-'
    c = ws.cell(row=r, column=10, value=p["dur"]); c.number_format = '0.0'
    ws.cell(row=r, column=11, value="SÍ" if p["main"] else "no")
    ws.cell(row=r, column=12, value=horario_semana(p))
    ws.cell(row=r, column=13, value=p["closed"] or "—")
    ws.cell(row=r, column=14, value=p["free"] or "—")
    ws.cell(row=r, column=15, value=p["book"] or "—")
    ws.cell(row=r, column=16, value=p["sub"])
    ws.cell(row=r, column=17, value=p["why"])
    ws.cell(row=r, column=18, value=p["addr"])
    c20 = ws.cell(row=r, column=20, value="Abrir ↗")
    c20.hyperlink = gmaps_url(p)
    c20.font = Font(name=F, size=10, color="0563C1", underline="single")
    _SRCN = {"ambas": "lista google", "lista": "lista google", "propuesta": "otros"}
    _SRCC = {"ambas": "5B3A91", "lista": "5B3A91", "propuesta": "7A736A"}
    _sv = p.get("src", "propuesta")
    c21 = ws.cell(row=r, column=21, value=_SRCN.get(_sv, _sv))
    c21.font = Font(name=F, size=10, bold=True, color=_SRCC.get(_sv, "000000"))
    if p.get("gname"):
        c21.comment = Comment("En la lista de Google figura como: " + p["gname"]
                              + "\n(coincide con lo que habíamos propuesto nosotros)", "Plan NYC")
    _mk = p.get("must") or 0
    c22 = ws.cell(row=r, column=22, value="SÍ" if _mk == 1 else "NO" if _mk == -1 else "")
    c22.fill = PatternFill("solid", fgColor=YELLOW if _mk != -1 else "EFEBE5")
    c22.font = Font(name=F, size=10, bold=True, color="8A6008" if _mk != -1 else "7B7268")
    c22.alignment = Alignment(horizontal="center")
    _dd = DAYMAP.get(p["id"], [])
    c19 = ws.cell(row=r, column=19, value=", ".join(_dd) if _dd else "—")
    if _dd:
        c19.font = Font(name=F, size=10, bold=True, color="1F6B3A")
        c19.fill = PatternFill("solid", fgColor=GREEN)
    if p["costN"]:
        ws.cell(row=r, column=9).comment = Comment(p["costN"], "Plan NYC")
    for col in range(1, 23):
        cc = ws.cell(row=r, column=col)
        if cc.font.name != F or not cc.font.size:
            cc.font = Font(name=F, size=10)
        if col in (20, 21, 22): continue
        cc.border = BOX
        cc.alignment = Alignment(wrap_text=(col in (12, 14, 15, 16, 17, 18, 19)), vertical="top",
                                 horizontal="center" if col in (6, 7, 8, 9, 10, 11) else "left")
    if p["cat"] in ("evento", "teatro"):
        for col in range(1, 23):
            _cc = ws.cell(row=r, column=col)
            if _cc.fill.start_color.rgb is None or _cc.fill.fill_type != "solid":
                _cc.fill = PatternFill("solid", fgColor=ORANGE)
        for col in (6, 7, 22):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=YELLOW)
    if p["free"]:
        ws.cell(row=r, column=14).fill = PatternFill("solid", fgColor=GREEN)
    ws.row_dimensions[r].height = 46
    r += 1
LAST = r - 1

dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False,
                    error="Solo 0, 1 o 2", errorTitle="Valor inválido",
                    prompt="0 = no me interesa · 1 = quizás · 2 = sí quiero", promptTitle="Nivel de interés")
ws.add_data_validation(dv); dv.add(f"F5:G{LAST}")
dvm = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True,
                     error="Escribí SÍ, NO, o dejalo vacío", errorTitle="Valor inválido",
                     prompt="SÍ = tiene que entrar sí o sí. NO = lo descartás a propósito. Vacío = sin decidir.",
                     promptTitle="Imprescindible")
ws.add_data_validation(dvm); dvm.add(f"V5:V{LAST}")
ws.auto_filter.ref = f"A4:V{LAST}"
ws.conditional_formatting.add(f"H5:H{LAST}",
    CellIsRule(operator="equal", formula=['"AMBOS"'], fill=PatternFill("solid", fgColor=GREEN)))

# ══════════════════════════ HOJA 3: ITINERARIO ══════════════════════════
ws = wb.create_sheet("Itinerario")
title(ws, "Itinerario día por día", "Editá 'Estado' y 'Mis notas'. Cada día tiene una alternativa al final.", span=7)
hdr(ws, 4, ["Día", "Hora", "Fin", "Quién", "Qué", "Lugar (ref. hoja Lugares)",
            "Traslado desde el anterior (estimado)", "Estado", "Mis notas"],
    [22, 8, 8, 10, 92, 28, 30, 13, 36])
ws.freeze_panes = "A5"
WHO = {"both": "Ambos", "jp": "Solo JP", "th": "Thais"}
pname = {p["id"]: p["n"] for p in places}
r = 5
for d in itin:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1,
        value=f"DÍA {d['n']} · {d['dow']} {d['date'][8:]}/{d['date'][5:7]} · {d['title']}  —  base: {d['base']}  ·  {WHO[d['who']]}")
    c.font = Font(name=F, bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22; r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value=d["note"])
    c.font = Font(name=F, size=9, italic=True, color="595959")
    c.fill = PatternFill("solid", fgColor=GREY)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 32; r += 1
    for bi, b in enumerate(d["blocks"]):
        nxt = d["blocks"][bi + 1]["t"] if bi + 1 < len(d["blocks"]) else ""
        ws.cell(row=r, column=2, value=b["t"]).font = Font(name=F, size=10, bold=True)
        ws.cell(row=r, column=3, value=nxt).font = Font(name=F, size=9, color="595959")
        ws.cell(row=r, column=4, value=WHO[b["who"]]).font = Font(name=F, size=9)
        cq = ws.cell(row=r, column=5, value=b["txt"]); cq.font = Font(name=F, size=10)
        cq.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=6, value=pname.get(b["pid"], "")).font = Font(name=F, size=9, color="595959")
        ct = ws.cell(row=r, column=7, value=b.get("trav", ""))
        ct.font = Font(name=F, size=9, italic=True, color="595959")
        ct.alignment = Alignment(wrap_text=True, vertical="top")
        ce = ws.cell(row=r, column=8, value=""); ce.fill = PatternFill("solid", fgColor=YELLOW)
        cn = ws.cell(row=r, column=9, value=""); cn.fill = PatternFill("solid", fgColor=YELLOW)
        if b["kind"] == "destacado":
            for col in range(2, 7):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=GREEN)
            cq.font = Font(name=F, size=10, bold=True)
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BOX
        ws.row_dimensions[r].height = 30 if len(b["txt"]) > 95 else 16
        r += 1
    if d["alt"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1, value="ALTERNATIVA / OJO:  " + d["alt"])
        c.font = Font(name=F, size=9, italic=True, color="833C00")
        c.fill = PatternFill("solid", fgColor=ORANGE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 32; r += 1
    r += 1

dv2 = DataValidation(type="list", formula1='"Pendiente,Reservado,Confirmado,Descartado"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"H5:H{r}")

# ══════════════════════════ HOJA 4: COSTOS ══════════════════════════
ws = wb.create_sheet("Costos")
title(ws, "Costos estimados", "Suma automática de lo marcado con interés = 2 en la hoja Lugares. Cambien allá y esto se actualiza.", span=6)
hdr(ws, 4, ["Concepto", "Juan Pablo", "Thais", "Total pareja", "Base del cálculo", ""],
    [42, 15, 15, 15, 70, 2])
r = 5
def crow(label, fjp, fth, base, bold=False, fill=None):
    global r
    c = ws.cell(row=r, column=1, value=label); c.font = Font(name=F, size=10, bold=bold)
    for col, f in ((2, fjp), (3, fth)):
        cc = ws.cell(row=r, column=col, value=f)
        cc.number_format = '$#,##0;($#,##0);-'; cc.font = Font(name=F, size=10, bold=bold)
    cc = ws.cell(row=r, column=4, value=f"=B{r}+C{r}")
    cc.number_format = '$#,##0;($#,##0);-'; cc.font = Font(name=F, size=10, bold=True)
    cb = ws.cell(row=r, column=5, value=base); cb.font = Font(name=F, size=9, color="595959")
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if fill:
        for col in range(1, 6): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)
    for col in range(1, 6): ws.cell(row=r, column=col).border = BOX
    ws.row_dimensions[r].height = 26
    r += 1

def atrac(col):
    """Entradas de atracciones: excluye Comida y Música, que se cuentan en sus propias filas."""
    return (f"=SUMIF(Lugares!${col}$5:${col}${LAST},2,Lugares!$I$5:$I${LAST})"
            f"-SUMIFS(Lugares!$I$5:$I${LAST},Lugares!${col}$5:${col}${LAST},2,Lugares!$C$5:$C${LAST},\"Comida\")"
            f"-SUMIFS(Lugares!$I$5:$I${LAST},Lugares!${col}$5:${col}${LAST},2,Lugares!$C$5:$C${LAST},\"Música\")")

crow("Entradas de atracciones (interés = 2)", atrac("F"), atrac("G"),
     "SUMIF sobre la hoja Lugares: suma 'US$ p/p' de las filas con interés 2, RESTANDO las categorías Comida y Música "
     "(esas se cuentan abajo en sus propias filas, para no contarlas dos veces). Si cambian un 0/1/2 allá, esto se recalcula.")
crow("Subte / bus (OMNY, tope semanal)", 70, 55,
     "OMNY 2026: $3,00 por viaje, tope $35 por período de 7 días. JP: 9 días en NYC = 2 períodos = $70 máximo. Thais: menos días de turismo, estimado $55.")
crow("NJ Transit Raritan ↔ NY Penn", 74, 37,
     "PLAN CONFIRMADO. JP: 4 tramos × $18,50 (vuelta a NJ el 31, ida el 1, vuelta el 2, ida el 3 con las maletas). "
     "Thais: 2 tramos (ida el 31, vuelta el 3). Los pasajes sueltos son lo más barato: no hay descuento por round-trip ni 10-trip. "
     "Desde julio de 2026 los one-way VENCEN a los 30 días de comprados.")
crow("Uber Branchburg ↔ estación Raritan", 72, 36,
     "JP 4 tramos y Thais 2, × ~$18 estimado (4 millas, zona semi-rural). El hotel NO tiene shuttle. "
     "Pedir el auto la noche anterior: a las 6 AM en Branchburg puede no haber conductores.")
crow("EWR ↔ Manhattan (ida y vuelta)", 38, 38,
     "Llegada: AirTrain + NJ Transit $17,25. Salida: subte $3,00 + NJ Transit $17,25 = $20,25.")
crow("Comida — TODAS las comidas del viaje", 630, 630,
     "9 días × ~$70 por persona por día, desayuno-almuerzo-cena incluidos. Realista si mezclan barrios étnicos "
     "($12-20 por comida: Jackson Heights, Sunset Park, Flushing, Arthur Avenue) con algunas cenas de $40-50. "
     "Esta fila reemplaza a los restaurantes individuales de la hoja Lugares — por eso se restan arriba.")
crow("Música en vivo (jazz y clubes)", 130, 90,
     "Village Vanguard ~$50 (cover + mínimo), Smalls/Mezzrow $25 con bebida, Bar LunÀtico $10 en efectivo, "
     "Dizzy's late night ~$20-30. Reemplaza a los clubes individuales de la hoja Lugares.")
crow("Broadway / teatro (opcional)", 65, 65,
     "Estimado con lotería o TKTS. Comprado directo, un musical grande promedia $92-112 por persona.")
r += 1
c = ws.cell(row=r, column=1, value="TOTAL ESTIMADO"); c.font = Font(name=F, size=12, bold=True, color="FFFFFF")
for col, f in ((2, f"=SUM(B5:B{r-2})"), (3, f"=SUM(C5:C{r-2})"), (4, f"=B{r}+C{r}")):
    cc = ws.cell(row=r, column=col, value=f)
    cc.number_format = '$#,##0'; cc.font = Font(name=F, size=12, bold=True, color="FFFFFF")
for col in range(1, 6):
    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=r, column=col).border = BOX
ws.cell(row=r, column=5, value="No incluye alojamiento ni vuelos.").font = Font(name=F, size=9, color="FFFFFF")
ws.row_dimensions[r].height = 24
r += 3
ws.cell(row=r, column=1, value="AHORRO DEL CALENDARIO GRATUITO").font = Font(name=F, size=11, bold=True, color=NAVY)
r += 1
hdr(ws, r, ["Día", "Qué es gratis", "Valor evitado (2 personas)", "", "", ""], [22, 60, 24, 2, 2, 2])
r += 1
for g in extras["passes"]["gratis_valor"]:
    ws.cell(row=r, column=1, value=g["d"]).font = Font(name=F, size=10, bold=True)
    ws.cell(row=r, column=2, value=g["q"]).font = Font(name=F, size=10)
    ws.cell(row=r, column=3, value=g["v"]).font = Font(name=F, size=10)
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=GREEN)
    r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(row=r, column=1, value=extras["passes"]["total_gratis"])
c.font = Font(name=F, size=10, bold=True, color="1F6F3D"); c.alignment = Alignment(wrap_text=True)
ws.row_dimensions[r].height = 28

# ══════════════════════════ HOJA 5: TRANSPORTE ══════════════════════════
ws = wb.create_sheet("Transporte")
title(ws, "Transporte", "Precios y horarios verificados en fuentes oficiales el 16 de agosto de 2026.", span=5)
ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 24; ws.column_dimensions["D"].width = 95
r = 4
def sect(t, sub=""):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=t)
    c.font = Font(name=F, bold=True, size=12, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[r].height = 22; r += 1
    if sub:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=sub); c.font = Font(name=F, size=9, italic=True, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="center"); ws.row_dimensions[r].height = 30; r += 1

def line(a, b_, c_, d_, fill=None, bold=False):
    global r
    for col, v in ((1, a), (2, b_), (3, c_), (4, d_)):
        cc = ws.cell(row=r, column=col, value=v)
        cc.font = Font(name=F, size=10, bold=bold)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        cc.border = BOX
        if fill: cc.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[r].height = 40 if len(str(d_)) > 100 else 26
    r += 1

for key in ("ewr_llegada", "ewr_salida"):
    t = extras["transport"][key]
    sect(t["titulo"], "RECOMENDADO: " + t["reco"])
    line("Opción", "Costo", "Tiempo", "Detalle", SKY, True)
    for o in t["opciones"]:
        line(o["m"], o["costo"], o["tiempo"], o["det"] + ("  ▸ " + o["total"] if o.get("total") else ""))
    if t.get("secuencia"): line("Secuencia realista", "", "", t["secuencia"], GREEN)
    line("⚠ Alerta", "", "", t["alerta"], ORANGE)
    r += 1

t = extras["transport"]["nj_nyc"]
sect(t["titulo"], "RECOMENDADO: " + t["reco"])
line("Hotel", "", "", t["hotel"], ORANGE)
line("Estación", "Distancia del hotel", "Estacionamiento", "Servicio", SKY, True)
for e in t["estaciones"]:
    line(e["e"], e["dist"], e["park"], e["serv"], GREEN if "✅" in e["e"] else None)
r += 1
line("Tipo de pasaje", "Precio", "", "Nota", SKY, True)
for f in t["tarifas"]:
    line(f["t"], f["p"], "", f["nota"], GREEN if "MÁS BARATA" in f["nota"] else None)
r += 1
line("Horarios", "", "", t["horarios"])
line("Trenes directos", "", "", t["oneseat"])
r += 1
line("Escenario", "Cálculo", "", "Total 4 días", SKY, True)
for cmp_ in t["comparativa"]:
    line(cmp_["esc"], cmp_["calc"], "", cmp_["tot"], GREEN if "✅" in cmp_["esc"] else (ORANGE if "❌" in cmp_["esc"] else None))
line("¿Conviene alquilar auto?", "NO", "", t["veredicto_auto"], ORANGE)
line("⚠ Riesgo", "", "", t["riesgo"], ORANGE)
r += 1

t = extras["transport"]["nyc_interno"]
sect(t["titulo"])
line("🔴 CAMBIO CLAVE 2026", "", "", t["cambio_clave"], ORANGE, True)
line("Tarifa", "$3,00", "", t["tarifa"])
line("Fare capping OMNY", "tope $35 / 7 días", "", t["capping"], GREEN)
line("⚠ CRÍTICO", "", "", t["critico"], ORANGE)
line("Cuenta del viaje", "", "", t["cuenta"])
for h in t["hoteles"]:
    line(h["h"], "", "", h["s"])

# ══════════════════════════ HOJA 6: PASES ══════════════════════════
ws = wb.create_sheet("Pases")
title(ws, "¿Conviene un pase?", "Análisis contra el itinerario propuesto.", span=5)
ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 26; ws.column_dimensions["D"].width = 60
ws.column_dimensions["E"].width = 70
r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value="VEREDICTO:  " + extras["passes"]["veredicto"])
c.font = Font(name=F, bold=True, size=13, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
c.alignment = Alignment(vertical="center", wrap_text=True, indent=1); ws.row_dimensions[r].height = 34
r += 2
ws.cell(row=r, column=1, value="POR QUÉ").font = Font(name=F, bold=True, size=11, color=NAVY); r += 1
for i, razon in enumerate(extras["passes"]["razones"], 1):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=f"{i}.  {razon}")
    c.font = Font(name=F, size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30; r += 1
r += 1
hdr(ws, r, ["Pase", "Precio adulto", "Validez", "Qué incluye", "Análisis para este viaje"],
    [34, 16, 26, 60, 70]); r += 1
for p in extras["passes"]["opciones"]:
    for col, v in ((1, p["p"]), (2, p["c"]), (3, p["v"]), (4, p["inc"]), (5, p["an"])):
        cc = ws.cell(row=r, column=col, value=v)
        cc.font = Font(name=F, size=10, bold=(col == 1))
        cc.alignment = Alignment(wrap_text=True, vertical="top"); cc.border = BOX
        if "🚫" in p["an"] or "❌" in p["an"]:
            cc.fill = PatternFill("solid", fgColor=ORANGE)
    ws.row_dimensions[r].height = 56; r += 1
r += 2
ws.cell(row=r, column=1, value="COMPRAR CON ANTICIPACIÓN").font = Font(name=F, bold=True, size=11, color=NAVY); r += 1
for x in extras["passes"]["comprar_ahora"]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="▸  " + x)
    c.font = Font(name=F, size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[r].height = 26; r += 1

# ══════════════════════════ HOJA 7: RESERVAS ══════════════════════════
ws = wb.create_sheet("Reservas")
title(ws, "Checklist de reservas", "Ordenado por urgencia. Marcá 'Hecho' en la columna amarilla.", span=5)
hdr(ws, 4, ["Hecho", "Qué reservar", "Cuándo", "Dónde", "Por qué"], [10, 44, 30, 34, 82])
ws.freeze_panes = "A5"
r = 5
for x in sorted(extras["reservas"], key=lambda z: -z["urg"]):
    cc = ws.cell(row=r, column=1, value=""); cc.fill = PatternFill("solid", fgColor=YELLOW)
    for col, v in ((2, x["q"]), (3, x["cuando"]), (4, x["url"]), (5, x["por"])):
        c2 = ws.cell(row=r, column=col, value=v)
        c2.font = Font(name=F, size=10, bold=(col == 2))
        c2.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BOX
        if x["urg"] == 3 and col > 1:
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ORANGE)
    ws.row_dimensions[r].height = 40; r += 1
dv3 = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True)
ws.add_data_validation(dv3); dv3.add(f"A5:A{r}")

# ══════════════════════════ HOJA 8: DECISIONES ══════════════════════════
ws = wb.create_sheet("Decisiones")
title(ws, "Decisiones abiertas", "Lo que falta definir, con los números para decidir.", span=4)
hdr(ws, 4, ["Decisión", "De qué se trata", "Los números", "Qué hace falta / mi recomendación"],
    [36, 70, 70, 70])
r = 5
for d in extras["decisiones"]:
    for col, v in ((1, d["t"]), (2, d["d"]), (3, d["num"]), (4, d["req"])):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name=F, size=10, bold=(col == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BOX
    ws.row_dimensions[r].height = 78; r += 1

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

wb.save(os.path.join(_R, "NYC_2026_Planificador.xlsx"))
print(f"OK - Excel con {len(wb.worksheets)} hojas, {LAST-4} lugares")
