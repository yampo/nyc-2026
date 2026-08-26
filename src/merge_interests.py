#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lee un Planificador editado por el usuario y extrae los intereses JP/Thais.

  python3 merge_interests.py <archivo.xlsx>

Busca las columnas POR NOMBRE de encabezado, no por posición, así funciona
con cualquier versión de la planilla (con o sin la columna "Tipo").
Escribe data/overrides.json = { "<id>": {"jp": n, "th": n, "must": bool} } y reporta el diff.
La columna IMPRESCINDIBLE (SÍ / vacío) se lee si está presente.
"""
import sys, json, unicodedata, re
from openpyxl import load_workbook
import os
_R = os.path.dirname(os.path.abspath(__file__))   # todo se resuelve relativo a este archivo, no a donde se corra

def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "", s.lower())

src = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/NYC_2026_Planificador_thais.xlsx"
places = json.load(open(os.path.join(_R, "data/places.json")))["places"]
by_norm = {norm(p["n"]): p for p in places}

wb = load_workbook(src, data_only=True)
if "Lugares" not in wb.sheetnames:
    sys.exit(f"ERROR: el archivo no tiene una hoja 'Lugares'. Tiene: {wb.sheetnames}")
ws = wb["Lugares"]

# localizar la fila de encabezados (la que contiene 'Lugar' y 'JP')
hdr_row = None
for r in range(1, 12):
    vals = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
    if "lugar" in vals and "jp" in vals:
        hdr_row = r; break
if not hdr_row:
    sys.exit("ERROR: no encontré la fila de encabezados con 'Lugar' y 'JP'.")

col = {}
for c in range(1, ws.max_column + 1):
    col[norm(ws.cell(row=hdr_row, column=c).value)] = c
for need in ("lugar", "jp", "thais"):
    if need not in col:
        sys.exit(f"ERROR: falta la columna '{need}'. Encontré: {sorted(k for k in col if k)}")

def as_int(v):
    if v is None or v == "": return None
    try:
        n = int(float(v))
        return n if n in (0, 1, 2) else None
    except (TypeError, ValueError):
        return None

ov, sin_match, invalidos, leidos = {}, [], [], 0
for r in range(hdr_row + 1, ws.max_row + 1):
    nombre = ws.cell(row=r, column=col["lugar"]).value
    if not nombre: continue
    p = by_norm.get(norm(nombre))
    if not p:
        sin_match.append(str(nombre)); continue
    jp, th = as_int(ws.cell(row=r, column=col["jp"]).value), as_int(ws.cell(row=r, column=col["thais"]).value)
    if jp is None or th is None:
        invalidos.append(f"{nombre} (JP={ws.cell(row=r,column=col['jp']).value!r}, "
                         f"Thais={ws.cell(row=r,column=col['thais']).value!r})")
        continue
    leidos += 1
    _m = ws.cell(row=r, column=col["imprescindible"]).value if "imprescindible" in col else None
    _m = str(_m).strip().upper() in ("SÍ", "SI", "SÍ ", "X", "TRUE", "1") if _m else False
    ov[p["id"]] = {"jp": jp, "th": th}
    if _m: ov[p["id"]]["must"] = True

# ── diff contra los valores por defecto ──────────────────────────────────
cambios = []
for p in places:
    o = ov.get(p["id"])
    if not o: continue
    if o["jp"] != p["jp"] or o["th"] != p["th"]:
        cambios.append((p, o))

no_leidos = [p["n"] for p in places if p["id"] not in ov]

print(f"Archivo:            {src}")
print(f"Fila de encabezado: {hdr_row}")
print(f"Filas leídas OK:    {leidos} de {len(places)} lugares")
print()
if sin_match:
    print(f"⚠️  {len(sin_match)} nombres del archivo NO coinciden con ningún lugar conocido:")
    for n in sin_match[:20]: print(f"     · {n}")
    if len(sin_match) > 20: print(f"     … y {len(sin_match)-20} más")
    print()
if invalidos:
    print(f"⚠️  {len(invalidos)} filas con valores fuera de 0/1/2 (se ignoran, queda el default):")
    for n in invalidos[:15]: print(f"     · {n}")
    print()
if no_leidos:
    print(f"ℹ️  {len(no_leidos)} lugares del catálogo no aparecieron en el archivo (mantienen el default):")
    for n in no_leidos[:15]: print(f"     · {n}")
    if len(no_leidos) > 15: print(f"     … y {len(no_leidos)-15} más")
    print()

print(f"═══ {len(cambios)} CAMBIOS DE INTERÉS ═══")
subidas_th = [x for x in cambios if x[1]["th"] > x[0]["th"]]
bajadas_th = [x for x in cambios if x[1]["th"] < x[0]["th"]]
subidas_jp = [x for x in cambios if x[1]["jp"] > x[0]["jp"]]
bajadas_jp = [x for x in cambios if x[1]["jp"] < x[0]["jp"]]
for etiqueta, grupo, quien in (("THAIS ↑", subidas_th, "th"), ("THAIS ↓", bajadas_th, "th"),
                               ("JP ↑", subidas_jp, "jp"), ("JP ↓", bajadas_jp, "jp")):
    if not grupo: continue
    print(f"\n{etiqueta}  ({len(grupo)})")
    for p, o in sorted(grupo, key=lambda x: x[0]["n"]):
        print(f"   {p['n'][:42]:44s} {p[quien]} → {o[quien]}")

json.dump(ov, open(os.path.join(_R, "data/overrides.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"\n✓ data/overrides.json escrito ({len(ov)} lugares)")
